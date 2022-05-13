from django.shortcuts import render

import xlsxwriter
import openpyxl

def excel_writer(request):
    filename="documents/LibroExcel2.xlsx"

    data=[
        ["Nombre","Apellido","Edad"],
        ["Jon","Snow",33],
        ["Daenerys","Targaryen",25],
        ["Tyrion","Lannister",40],
        ["Jaime","Lannister",35],
        ["Cersei","Lannister",36]
    ]

    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet("Hoja 1")

    for row in range(len(data)):
        for col in range(len(data[row])):
            worksheet.write(row, col, data[row][col])

    workbook.close()

    return render(request, 'csv.html')

def excel_read(request):
    filename="documents/LibroExcel2.xlsx"
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, 3):
            print(col[i].value, end="\t")
        print("")
    
    return render(request, 'csv.html')

